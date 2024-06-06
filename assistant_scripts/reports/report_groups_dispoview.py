import os, datetime

import pandas as pd
import openpyxl.utils

from openpyxl import load_workbook
from openpyxl.styles import PatternFill, NamedStyle, Font, Border, Side
from openpyxl.formatting.rule import CellIsRule
from assistant_scripts.read_data.read_dispoview_data import DispoviewDataReader


class GroupsDispoview:
    def __init__(
        self,
        dispo_file_path: str,
        groups_file_path: str,
        supply_file_path: str,
    ):
        self.dispo_file_path = dispo_file_path
        self.groups_file_path = groups_file_path
        self.supply_file_path = supply_file_path

        self.raw_dispoview = None

        self.raw_groups = None
        self.ready_groups = None

        self.unique_groups = None
        self.unique_codenumbers = None
        self.component_to_codenumber = None

        self.supply_confirmed = None
        self.supply_requested = None

        self.groups_balances = None
        self.weeks = None

    def _read_dispoview(self):
        dispoview = DispoviewDataReader(dispo_file_path=self.dispo_file_path)
        dispoview()
        self.raw_dispoview = dispoview.ready_dispoview

    def _read_groups(self):
        self.raw_groups = pd.read_excel(self.groups_file_path)

    def _select_groups(self):
        self.ready_groups = self.raw_groups[
            ["GROUP", "COMPONENT", "CODENUMBER", "GROUP_DESCRIPTION", "GROUP_STATUS"]
        ]
        self.ready_groups = self.ready_groups.drop_duplicates(inplace=False)
        self.ready_groups.reset_index(inplace=True, drop=True)
        self.unique_groups = sorted(self.ready_groups["GROUP"].unique())
        self.unique_codenumbers = sorted(self.ready_groups["CODENUMBER"].unique())
        self.component_to_codenumber = dict(
            zip(self.ready_groups["COMPONENT"], self.ready_groups["CODENUMBER"])
        )

    def _format_week_year(self, date):
        week_number = date.isocalendar().week
        year = date.year
        return f"W{week_number}.{year}"

    def _read_supply(self):
        def process_supply_sheet(sheet_name):
            df = pd.read_excel(self.supply_file_path, sheet_name=sheet_name)
            df["ETD_DATE_WEEK"] = df["ETD_DATE_WEEK"].apply(self._format_week_year)
            df["CODENUMBER"] = df["COMPONENT"].map(self.component_to_codenumber)
            return df

        self.supply_confirmed = process_supply_sheet("supply_confirmed")
        self.supply_requested = process_supply_sheet("supply_requested")

    def _mergre_groups_dispoview(self):
        self.all_merged_data = pd.merge(
            self.ready_groups[["CODENUMBER", "GROUP", "GROUP_DESCRIPTION"]],
            self.raw_dispoview,
            on="CODENUMBER",
            how="left",
        )
        self.weeks = list(self.all_merged_data.columns[4:])

    def _formula_column(self, row, first: bool):
        if row["DATA"] == "Forecast_confirmed":
            if first:
                part_a = f'_xlfn.SUMIFS(All_data!E:E,All_data!$B:$B,$A{row.name +2},All_data!$D:$D,"Stock")'
                sub_part_a = f"E$1"
                sub_part_b = f"All_data!E:E"
            else:
                part_a = f"E{row.name +2}"
                sub_part_a = f"F$1"
                sub_part_b = f"All_data!F:F"
            part_b = f'_xlfn.SUMIFS({sub_part_b},All_data!$B:$B,$A{row.name +2},All_data!$D:$D,"NetForecast")'
            part_c = f"_xlfn.SUMIFS(Supply_confirmed!$E:$E,Supply_confirmed!$A:$A,$A{row.name +2},Supply_confirmed!$D:$D,{sub_part_a})"
            return f"={part_a}-{part_b}+{part_c}"

        if row["DATA"] == "Orders_confirmed":
            if first:
                part_a = f'_xlfn.SUMIFS(All_data!E:E,All_data!$B:$B,$A{row.name +2},All_data!$D:$D,"Stock")'
                sub_part_a = f"E$1"
                sub_part_b = f"All_data!E:E"
            else:
                part_a = f"E{row.name +2}"
                sub_part_a = f"F$1"
                sub_part_b = f"All_data!F:F"
            part_b = f'_xlfn.SUMIFS({sub_part_b},All_data!$B:$B,$A{row.name +2},All_data!$D:$D,"CustOrders")'
            part_c = f"_xlfn.SUMIFS(Supply_confirmed!$E:$E,Supply_confirmed!$A:$A,$A{row.name +2},Supply_confirmed!$D:$D,{sub_part_a})"
            return f"={part_a}-{part_b}+{part_c}"

    def _create_groups_balances(self):
        main_headers = ["GROUP", "GROUP_DESCRIPTION", "DATA", "COMMENTS"] + self.weeks
        groups_descriptions = (
            self.ready_groups[["GROUP", "GROUP_DESCRIPTION"]]
            .drop_duplicates()
            .reset_index(drop=True)
        )
        self.groups_balances = pd.DataFrame(columns=main_headers)

        forecast_confirmed = groups_descriptions.copy()
        forecast_confirmed["DATA"] = "Forecast_confirmed"

        orders_confirmed = groups_descriptions.copy()
        orders_confirmed["DATA"] = "Orders_confirmed"

        forecast_requested = groups_descriptions.copy()
        forecast_requested["DATA"] = "Forecast_requested"

        healthy_stock_forecast = groups_descriptions.copy()
        healthy_stock_forecast["DATA"] = "Healthy_stock_forecast"

        self.groups_balances = pd.concat(
            [
                self.groups_balances,
                forecast_confirmed,
                forecast_requested,
                healthy_stock_forecast,
                orders_confirmed,
            ],
            ignore_index=True,
        )
        self.groups_balances.iloc[:, 4] = self.groups_balances.apply(
            lambda row: self._formula_column(row, True), axis=1
        )
        self.groups_balances.iloc[:, 5] = self.groups_balances.apply(
            lambda row: self._formula_column(row, False), axis=1
        )

    def _apply_excel_formatting(self, final_file_path):
        redFill = PatternFill(
            start_color="FF7276", end_color="FF7276", fill_type="solid"
        )
        rule_negative = CellIsRule(
            operator="lessThan", formula=[0], stopIfTrue=True, fill=redFill
        )
        workbook = load_workbook(filename=final_file_path)
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            sheet.conditional_formatting.add("B2:Z1000", rule_negative)
        workbook.save(final_file_path)

    def _save_to_excel(self):
        now = datetime.datetime.now()
        filename = f"Report_groups_dispoview_{now.strftime('%d%m%Y_%H%M')}.xlsx"
        directory_path = os.path.dirname(self.dispo_file_path)
        report_file_path = os.path.join(directory_path, filename)

        writer = pd.ExcelWriter(report_file_path)
        self.groups_balances.to_excel(
            writer, sheet_name=f"Groups_balances", index=False
        )
        self.all_merged_data.to_excel(writer, sheet_name=f"All_data", index=False)
        self.supply_confirmed.to_excel(
            writer, sheet_name=f"Supply_confirmed", index=False
        )
        self.supply_confirmed.to_excel(
            writer, sheet_name=f"Supply_requested", index=False
        )
        self.raw_groups.to_excel(writer, sheet_name=f"Groups", index=False)
        writer._save()
        self._apply_excel_formatting(report_file_path)

    def __call__(self):
        self._read_dispoview()
        self._read_groups()
        self._select_groups()
        self._read_supply()
        self._mergre_groups_dispoview()
        self._create_groups_balances()
        self._save_to_excel()
