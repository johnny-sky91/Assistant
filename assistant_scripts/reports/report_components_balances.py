import os
import pandas as pd
from datetime import datetime, timedelta
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, NamedStyle, Font, Border, Side
from openpyxl.formatting.rule import CellIsRule
from assistant_scripts.read_data.read_components_data import ComponentsDataReader
import openpyxl.utils


class ComponentsBalances:
    def __init__(self, path_groups: str, path_components: str, path_supply: str):
        self.path_groups = path_groups
        self.path_components = path_components
        self.path_supply = path_supply
        self.groups = None
        self.supply = None
        self.orders = None
        self.demand = None
        self.stock = None
        self.groups_balances = None
        self.weeks = None
        self.report_sources = None

    def create_weeks(self):
        actual_date = datetime.now()
        how_many_weeks = 16
        next_monday = (
            actual_date - timedelta(days=actual_date.weekday()) + timedelta(weeks=1)
        ).replace(hour=0, minute=0, second=0, microsecond=0)

        first_monday = (actual_date - timedelta(days=actual_date.weekday())).replace(
            hour=0, minute=0, second=0, microsecond=0
        )
        self.weeks = [first_monday] + [
            next_monday - timedelta(weeks=i)
            for i in range(0, -(how_many_weeks - 1), -1)
        ]

    def get_groups(self):
        self.groups = pd.read_excel(self.path_groups, sheet_name="groups")
        self.groups = self.groups[
            ["COMPONENT", "GROUP", "GROUP_DESCRIPTION"]
        ].drop_duplicates()
        self.groups.reset_index(drop=True, inplace=True)

    def get_components_data(self):
        comp = ComponentsDataReader(
            file_path=self.path_components,
            fix_weeks=True,
            add_forecast=False,
            to_excel=False,
        )
        comp_data = comp()
        self.orders = comp_data.orders
        self.demand = comp_data.demand_plan
        self.stock = comp_data.components_stock

    def get_supply(self):
        self.supply = pd.read_excel(self.path_supply, sheet_name="current_info")

    def add_groups(self):
        to_add_groups = [self.orders, self.demand, self.stock]
        for data in to_add_groups:
            data["GROUP"] = data.apply(
                lambda row: f'=_xlfn.XLOOKUP(B{row.name + 2},Components_groups!A:A,Components_groups!B:B,"NOT_FOUND")',
                axis=1,
            )

            data.insert(0, "GROUP", data.pop("GROUP"))

    def prepare_groups_balances(self):
        groups_balances_headers = [
            "Category",
            "Description",
            "Group",
            "Data",
            "Current stock",
            "Healthy Stock Level vs. Forecast",
            "Healthy Stock Level vs. Actuals",
            "Comments",
        ] + self.weeks
        self.groups_balances = pd.DataFrame(columns=groups_balances_headers)
        groups = pd.DataFrame(
            self.groups[["GROUP", "GROUP_DESCRIPTION"]].drop_duplicates()
        )
        self.groups_balances["Description"] = 2 * groups["GROUP_DESCRIPTION"].to_list()
        self.groups_balances["Group"] = 2 * groups["GROUP"].to_list()
        midpoint = len(self.groups_balances) // 2
        self.groups_balances["Data"].loc[:midpoint] = "Stock vs. Real customer orders"
        self.groups_balances["Data"].loc[midpoint:] = "Stock vs. Forecast"
        self.groups_balances["Category"] = "MEM"

    def _formula_first_column(self, row):
        part_a = f"E{row.name +2}"
        part_b = (
            f"_xlfn.SUMIFS(Supply!$E:$E,Supply!$D:$D,I$1,Supply!$A:$A,$C{row.name +2})"
        )
        part_c = f'_xlfn.IF($D{row.name +2}="Stock vs. Real customer orders",SUMIFS(Order!C:C,Order!$A:$A,$C{row.name +2}),SUMIFS(Demand_Plan!C:C,Demand_Plan!$A:$A,$C{row.name +2}))'
        return f"={part_a}+{part_b}-{part_c}"

    def _formula_next_columns(self, row):
        part_a = f"I{row.name +2}"
        part_b = (
            f"_xlfn.SUMIFS(Supply!$E:$E,Supply!$D:$D,J$1,Supply!$A:$A,$C{row.name +2})"
        )
        part_c = f'_xlfn.IF($D{row.name +2}="Stock vs. Real customer orders",SUMIFS(Order!D:D,Order!$A:$A,$C{row.name +2}),SUMIFS(Demand_Plan!D:D,Demand_Plan!$A:$A,$C{row.name +2}))'
        return f"={part_a}+{part_b}-{part_c}"

    def _formula_stock_columns(self, row):
        return f"=_xlfn.SUMIFS(Stock!$E:$E,Stock!$A:$A,C{row.name +2})"

    def _formula_avg_demand_columns(self, row):
        return f"=_xlfn.AVERAGE(C{row.name +2}:V{row.name +2})*4"

    def _formula_dos_forecast(self, row):
        part_a = f"_xlfn.SUMIFS(Demand_Plan!$W:$W,Demand_Plan!$A:$A,$C{row.name +2})"
        part_b = 1.5
        return f"={part_a}*{part_b}"

    def _apply_formulas(self):
        self.groups_balances.iloc[:, 8] = self.groups_balances.apply(
            lambda row: self._formula_first_column(row), axis=1
        )
        self.groups_balances.iloc[:, 9] = self.groups_balances.apply(
            lambda row: self._formula_next_columns(row), axis=1
        )
        self.groups_balances.iloc[:, 4] = self.groups_balances.apply(
            lambda row: self._formula_stock_columns(row), axis=1
        )
        self.groups_balances.iloc[:, 5] = self.groups_balances.apply(
            lambda row: self._formula_dos_forecast(row), axis=1
        )
        self.demand["AVG"] = self.demand.apply(
            lambda row: self._formula_avg_demand_columns(row), axis=1
        )

    def _apply_excel_formatting(self, final_file_path):
        redFill = PatternFill(
            start_color="FF7276", end_color="FF7276", fill_type="solid"
        )
        rule_negative = CellIsRule(
            operator="lessThan", formula=[0], stopIfTrue=True, fill=redFill
        )
        workbook = load_workbook(filename=final_file_path)
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            sheet.conditional_formatting.add("B2:AE1000", rule_negative)

        sheet = workbook["Groups_balances"]
        columns_to_format = list(range(9, 25))
        date_format = NamedStyle(name="date_format", number_format="m/d")
        date_format.font = Font(bold=True)
        date_format.border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )
        for column_number in columns_to_format:
            column_letter = openpyxl.utils.get_column_letter(column_number)
            cell = sheet[column_letter + "1"]
            cell.style = date_format
        workbook.save(final_file_path)

    def get_report_sources(self):
        source_files = [self.path_groups, self.path_components, self.path_supply]
        source_files = [os.path.basename(x) for x in source_files]
        data_types = ["Groups", "Components", "Supply"]
        self.report_sources = pd.DataFrame(
            {
                "Data_type": data_types,
                "Source_file": source_files,
            }
        )

    def save_to_excel(self):
        now = datetime.now()
        filename = f"Report_groups_balances_{now.strftime('%d%m%Y_%H%M')}.xlsx"
        directory_path = os.path.dirname(self.path_components)
        report_file_path = os.path.join(directory_path, filename)

        writer = pd.ExcelWriter(report_file_path)
        self.report_sources.to_excel(writer, sheet_name=f"INFO", index=False)
        self.groups_balances.to_excel(
            writer, sheet_name=f"Groups_balances", index=False
        )
        self.groups.to_excel(writer, sheet_name=f"Components_groups", index=False)
        self.demand.to_excel(writer, sheet_name=f"Demand_Plan", index=False)
        self.orders.to_excel(writer, sheet_name=f"Order", index=False)
        self.stock.to_excel(writer, sheet_name=f"Stock", index=False)
        self.supply.to_excel(writer, sheet_name=f"Supply", index=False)

        writer._save()
        self._apply_excel_formatting(report_file_path)

    def __call__(self):
        self.get_groups()
        self.create_weeks()
        self.get_components_data()
        self.get_supply()
        self.add_groups()
        self.prepare_groups_balances()
        self._apply_formulas()
        self.get_report_sources()
        self.save_to_excel()
