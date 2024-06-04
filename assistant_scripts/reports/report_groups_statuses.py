import os, datetime
import pandas as pd

from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.formatting.rule import CellIsRule

from assistant_scripts.read_data.read_components_data import ComponentsDataReader
from assistant_scripts.read_data.read_products_data import ProductsDataReader
from assistant_scripts.read_data.read_companion_data import CompanionDbReader


class GroupsStatuses:
    def __init__(
        self,
        path_components: str,
        path_products: str,
        path_groups: str,
        path_supply: str,
        path_db: str,
    ):
        self.path_components = path_components
        self.path_products = path_products
        self.path_groups = path_groups
        self.path_supply = path_supply
        self.path_db = path_db
        self.groups = None
        self.all_products_orders = None
        self.all_products_supply = None
        self.all_products_balances = None
        self.all_components_stock = None
        self.all_components_supply = None
        self.groups_summary = None
        self.unique_groups = None
        self.db_info = None
        self.report_file_path = None

    def get_components_data(self):
        comp = ComponentsDataReader(
            file_path=self.path_components,
            fix_weeks=True,
            add_forecast=False,
            to_excel=False,
        )
        comp()
        self.all_components_stock = comp.components_stock

        self.all_components_stock = pd.merge(
            self.all_components_stock,
            self.groups[["GROUP", "COMPONENT"]],
            on="COMPONENT",
            how="left",
        )
        self.all_components_stock.insert(
            0, "GROUP", self.all_components_stock.pop("GROUP")
        )

        self.all_components_stock = self.all_components_stock.drop_duplicates()

    def get_products_data(self):
        products = ProductsDataReader(
            file_path=None, folder_path=self.path_products, to_excel=False
        )
        products()
        self.all_products_orders = products.orders
        self.all_products_supply = products.supply
        self.all_products_balances = products.balances
        self._update_all_products_orders()
        self._update_all_products_supply()
        self._update_all_products_balances()

    def _update_all_products_orders(self):
        self.all_products_orders = pd.merge(
            self.all_products_orders,
            self.groups[["GROUP", "SOI"]],
            on="SOI",
            how="left",
        )
        self.all_products_orders.insert(
            0, "GROUP", self.all_products_orders.pop("GROUP")
        )
        self.all_products_orders["ORDERS_CUMULATIVE"] = self.all_products_orders.iloc[
            :, 3:
        ].sum(axis=1)
        self.all_products_orders.insert(
            2, "ORDERS_CUMULATIVE", self.all_products_orders.pop("ORDERS_CUMULATIVE")
        )
        self.all_products_orders = self.all_products_orders.drop_duplicates()

    def _update_all_products_supply(self):
        self.all_products_supply = pd.merge(
            self.all_products_supply,
            self.groups[["GROUP", "SOI"]],
            on="SOI",
            how="left",
        )
        self.all_products_supply.insert(
            0, "GROUP", self.all_products_supply.pop("GROUP")
        )
        self.all_products_supply["SUPPLY_CUMULATIVE"] = self.all_products_supply.iloc[
            :, 2:
        ].sum(axis=1)
        self.all_products_supply.insert(
            2, "SUPPLY_CUMULATIVE", self.all_products_supply.pop("SUPPLY_CUMULATIVE")
        )
        self.all_products_supply = self.all_products_supply.drop_duplicates()

    def _update_all_products_balances(self):
        self.all_products_balances = pd.merge(
            self.all_products_balances,
            self.groups[["GROUP", "SOI"]],
            on="SOI",
            how="left",
        )
        self.all_products_balances.insert(
            0, "GROUP", self.all_products_balances.pop("GROUP")
        )
        self.all_products_balances["BALANCE_CUMULATIVE"] = (
            self.all_products_balances.iloc[:, -1]
        )
        self.all_products_balances.insert(
            2,
            "BALANCE_CUMULATIVE",
            self.all_products_balances.pop("BALANCE_CUMULATIVE"),
        )
        self.all_products_balances = self.all_products_balances.drop_duplicates()

    def get_groups_data(self):
        self.groups = pd.read_excel(self.path_groups, sheet_name="groups")
        self.unique_groups = self.groups["GROUP"].unique()

    def get_supply_data(self):
        self.all_components_supply = pd.read_excel(
            self.path_supply, sheet_name="supply_confirmed"
        )

    def get_db_info_data(self):
        self.db_info = CompanionDbReader(path_db=self.path_db)
        self.db_info()

    def get_groups_summary(self):
        self.groups_summary = pd.DataFrame(index=self.unique_groups, columns=[])

        groups_statuses = self.groups.copy()
        groups_statuses = groups_statuses[["GROUP", "GROUP_STATUS"]]
        groups_statuses.drop_duplicates(inplace=True)
        groups_statuses.reset_index(inplace=True, drop=True)
        groups_statuses.set_index("GROUP", inplace=True)

        products_orders = self.all_products_orders.copy()
        products_orders = products_orders[["GROUP", "ORDERS_CUMULATIVE"]]
        products_orders = (
            products_orders.groupby("GROUP")["ORDERS_CUMULATIVE"].sum().reset_index()
        )
        products_orders.set_index("GROUP", inplace=True)

        products_supply = self.all_products_supply.copy()
        products_supply = products_supply[["GROUP", "SUPPLY_CUMULATIVE"]]
        products_supply = (
            products_supply.groupby("GROUP")["SUPPLY_CUMULATIVE"].sum().reset_index()
        )
        products_supply.set_index("GROUP", inplace=True)

        products_balances = self.all_products_balances.copy()
        products_balances = products_balances[["GROUP", "BALANCE_CUMULATIVE"]]
        products_balances = (
            products_balances.groupby("GROUP")["BALANCE_CUMULATIVE"].min().reset_index()
        )
        products_balances.set_index("GROUP", inplace=True)

        components_stock = self.all_components_stock.copy()
        components_stock = components_stock[["GROUP", "TOTAL_STOCK"]]
        components_stock = (
            components_stock.groupby("GROUP")["TOTAL_STOCK"].sum().reset_index()
        )
        components_stock.set_index("GROUP", inplace=True)

        list_summary = [
            self.groups_summary,
            groups_statuses,
            products_orders,
            products_supply,
            products_balances,
            components_stock,
            # components_supply,
        ]
        self.groups_summary = pd.concat(
            list_summary,
            axis=1,
        )
        self.groups_summary.index.name = "GROUP"
        self.groups_summary.fillna(0, inplace=True)
        self.groups_summary = self.groups_summary.rename(
            index=lambda x: f'=HYPERLINK("#{x}!A1","{x}")'
        )
        self.groups_summary.rename(
            {"BALANCE_CUMULATIVE": "MIN_BALANCE_CUMULATIVE"}, inplace=True
        )

    def _one_group_info(self, group):
        group_info = pd.DataFrame(
            self.db_info.groups[self.db_info.groups["GROUP"] == group]
        )
        sheets = ["SUMMARY", "ALL_SOI_SUPPLY", "ALL_SOI_BALANCES", "ALL_SOI_ORDERS"]
        for sheet in sheets:
            group_info[sheet] = f'=HYPERLINK("#{sheet}!A1", "{sheet}")'
        group_info.set_index("GROUP", inplace=True)

        return group_info

    def _one_group_matrix(self, group):
        matrix = pd.DataFrame(self.groups[self.groups["GROUP"] == group])
        data = matrix[["SOI", "COMPONENT", "USAGE"]]
        group_matrix = data.pivot_table(
            index="SOI",
            columns="COMPONENT",
            values="USAGE",
            aggfunc="first",
            fill_value=0,
        )
        group_products_notes = self.db_info.products[
            self.db_info.products["GROUP"] == group
        ]
        group_products_notes.set_index("SOI", inplace=True)

        group_matrix = group_matrix.reindex(sorted(group_matrix.columns), axis=1)
        group_matrix = pd.concat([group_matrix, group_products_notes], axis=1)
        group_matrix = group_matrix.rename_axis("SOI")
        group_matrix.drop(["GROUP"], axis=1, inplace=True)

        return group_matrix

    def _one_group_components_stock(self, group):
        components_info = pd.DataFrame(
            self.db_info.components[self.db_info.components["GROUP"] == group]
        )
        components_info.set_index("COMPONENT", inplace=True)
        components_info.drop(["GROUP"], axis=1, inplace=True)
        group_stock = pd.DataFrame(
            self.all_components_stock[self.all_components_stock["GROUP"] == group]
        )
        group_stock.set_index("COMPONENT", inplace=True)
        components_info = pd.concat([components_info, group_stock], axis=1)
        components_info.drop(["GROUP"], axis=1, inplace=True)
        components_info["TOTAL_STOCK"] = None
        components_info.iloc[0, components_info.columns.get_loc("TOTAL_STOCK")] = (
            components_info["WAREHOUSE_STOCK"].sum()
            + components_info["FACTORY_STOCK"].sum()
        )
        components_info = components_info.sort_index()

        return components_info

    def _one_group_components_supply(self, group):
        group_components_supply = pd.DataFrame(
            self.all_components_supply[self.all_components_supply["GROUP"] == group]
        )
        group_components_supply.drop(["GROUP"], axis=1, inplace=True)
        group_components_supply.set_index("COMPONENT", inplace=True)
        group_components_supply = group_components_supply.sort_index()

        return group_components_supply

    def _one_group_products_supply(self, group):
        group_products_supply = pd.DataFrame(
            self.all_products_supply[self.all_products_supply["GROUP"] == group]
        )
        group_products_supply.set_index("SOI", inplace=True)
        group_products_supply.drop(["GROUP"], axis=1, inplace=True)
        group_products_supply = group_products_supply.loc[
            :, group_products_supply.sum() != 0
        ]
        group_products_supply = group_products_supply.sort_index()

        return group_products_supply

    def _one_group_products_balances(self, group):
        group_products_balances = pd.DataFrame(
            self.all_products_balances[self.all_products_balances["GROUP"] == group]
        )
        group_products_balances.set_index("SOI", inplace=True)
        group_products_balances.drop(["GROUP"], axis=1, inplace=True)
        group_products_balances = group_products_balances.loc[
            :, group_products_balances.sum() != 0
        ]
        group_products_balances = group_products_balances.sort_index()
        return group_products_balances

    def _one_group_products_orders(self, group):
        group_products_orders = pd.DataFrame(
            self.all_products_orders[self.all_products_orders["GROUP"] == group]
        )
        group_products_orders.set_index("SOI", inplace=True)
        group_products_orders.drop(["GROUP"], axis=1, inplace=True)
        group_products_orders = group_products_orders.loc[
            :, group_products_orders.sum() != 0
        ]
        group_products_orders = group_products_orders.sort_index()

        return group_products_orders

    def prepare_one_group_data(self, group):
        all_data_group = []

        group_info = self._one_group_info(group=group)
        all_data_group.append(group_info)

        group_matix = self._one_group_matrix(group=group)
        all_data_group.append(group_matix)

        group_component_info = self._one_group_components_stock(group=group)
        all_data_group.append(group_component_info)

        group_components_supply = self._one_group_components_supply(group=group)
        all_data_group.append(group_components_supply)

        group_products_supply = self._one_group_products_supply(group=group)
        all_data_group.append(group_products_supply)

        group_products_balances = self._one_group_products_balances(group=group)
        all_data_group.append(group_products_balances)

        group_products_orders = self._one_group_products_orders(group=group)
        all_data_group.append(group_products_orders)

        positions = [0]
        for i, df in enumerate(all_data_group[:-1]):
            positions.append(len(df) + positions[i] + 2)
        return positions, all_data_group

    def save_to_excel(self):
        now = datetime.datetime.now()
        filename = f"Report_groups_statuses_{now.strftime('%d%m%Y_%H%M')}.xlsx"
        directory_path = os.path.dirname(self.path_components)
        self.report_file_path = os.path.join(directory_path, filename)
        writer = pd.ExcelWriter(self.report_file_path)

        self.groups_summary.to_excel(writer, sheet_name=f"SUMMARY", index=True)
        self.groups.to_excel(writer, sheet_name=f"GROUPS", index=True)

        self.all_products_orders.to_excel(
            writer, sheet_name=f"ALL_SOI_ORDERS", index=False
        )
        self.all_products_supply.to_excel(
            writer, sheet_name=f"ALL_SOI_SUPPLY", index=False
        )
        self.all_products_balances.to_excel(
            writer, sheet_name=f"ALL_SOI_BALANCES", index=False
        )
        self.all_components_stock.to_excel(
            writer, sheet_name=f"ALL_COMPONENTS_STOCK", index=False
        )
        self.all_components_supply.to_excel(
            writer, sheet_name=f"ALL_COMPONENTS_SUPPLY", index=False
        )

        for group in self.unique_groups:
            ready_sheet = self.prepare_one_group_data(group)
            for position, data in zip(ready_sheet[0], ready_sheet[1]):
                data.to_excel(
                    writer,
                    sheet_name=f"{group}",
                    startrow=position,
                    index=True,
                )

        writer._save()

    def _apply_excel_formatting(self):
        redFill = PatternFill(
            start_color="FF7276", end_color="FF7276", fill_type="solid"
        )
        rule_zero_less = CellIsRule(
            operator="lessThan", formula=[0], stopIfTrue=True, fill=redFill
        )
        rule_equal = CellIsRule(
            operator="equal", formula=['"Not active - EOL"'], fill=redFill
        )
        workbook = load_workbook(filename=self.report_file_path)
        blank_sheet_position = workbook.worksheets.index(workbook["SUMMARY"])
        sheets = workbook._sheets.copy()
        sheets.insert(0, sheets.pop(blank_sheet_position))
        workbook._sheets = sheets

        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            sheet.conditional_formatting.add("B2:CT1000", rule_zero_less)
            sheet.conditional_formatting.add("B2:J100", rule_equal)
            for column in range(ord("A"), ord("J") + 1):
                column_letter = chr(column)
                sheet.column_dimensions[column_letter].width = 18

        workbook.save(self.report_file_path)

    def __call__(self):
        self.get_groups_data()
        self.get_db_info_data()
        self.get_components_data()
        self.get_products_data()
        self.get_supply_data()
        self.get_groups_summary()
        self.save_to_excel()
        self._apply_excel_formatting()
